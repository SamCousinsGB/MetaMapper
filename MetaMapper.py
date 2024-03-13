import os
import shutil
import argparse
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter

NAMESPACE = "{http://soap.sforce.com/2006/04/metadata}"

COLUMN_HEADINGS = {
    'object': 'Object',
    'fullName': 'API Name',
    'label': 'Label',
    'deprecated': 'Deprecated',
    'externalId': 'External ID',
    'length': 'Length',
    'required': 'Required',
    'type': 'Field Type',
    'unique': 'Unique',
    'referenceTo': 'Reference To',
    'relationshipName': 'Relationship Name',
    'picklistValues': 'Picklist Values',
    'picklistType': 'Picklist Type',
    'formula': 'Formula',
    'formulaTreatBlanksAs': 'Formula Treat Blanks As',
    'trackFeedHistory': 'Track Feed History',
    'trackHistory': 'Track History'
}


def parse_xml(xml_file, field_type='all'):

    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        object_name = os.path.basename(os.path.dirname(os.path.dirname(xml_file)))
        field_name = root.find(f'.//{NAMESPACE}fullName').text
        # Determine if the field is custom based on '__c' in the field name
        is_custom_field = '__c' in field_name
        
        # Determine if the field should be included based on the field_type argument
        if field_type == 'custom' and not is_custom_field:
            return None  # Skip non-custom fields when filtering for custom
        elif field_type == 'standard' and is_custom_field:
            return None  # Skip custom fields when filtering for standard
        
        field_info = {'Object': object_name, 'API Name': field_name}
        for child in root:
            tag = child.tag.split('}')[-1]
            if tag in COLUMN_HEADINGS:
                if tag == 'type' and child.text == 'Picklist':
                    values = root.find(f".//{NAMESPACE}valueSet/{NAMESPACE}valueSetDefinition")
                    if values is not None:
                        picklist_values = '; '.join(v.find(f"{NAMESPACE}fullName").text for v in values if v.find(f"{NAMESPACE}fullName") is not None)
                        field_info['Picklist Values'] = picklist_values
                else:
                    field_info[COLUMN_HEADINGS[tag]] = child.text
        # Extract trackFeedHistory and trackHistory values
        trackFeedHistory = root.find(f'.//{NAMESPACE}trackFeedHistory')
        field_info['Track Feed History'] = trackFeedHistory.text if trackFeedHistory is not None else 'false'
        trackHistory = root.find(f'.//{NAMESPACE}trackHistory')
        field_info['Track History'] = trackHistory.text if trackHistory is not None else 'false'

        return field_info
    except Exception as e:
        print(f"Error parsing XML file '{xml_file}': {e}")
        return None





def find_xml_files():
    """
    Find all XML files representing fields in both 'force-app' and 'unpackaged' directories
    based on folder structure from the current working directory.
    """
    xml_files = []
    root_dir = os.getcwd()  # Use the current working directory as the root
    for base_dir in ['force-app', 'unpackaged']:
        sfdx_objects_folder = os.path.join(root_dir, base_dir, 'main', 'default', 'objects')
        if os.path.exists(sfdx_objects_folder):
            for root, dirs, files in os.walk(sfdx_objects_folder):
                for file in files:
                    if file.endswith('.field-meta.xml'):
                        print(f"Found XML file for field: {file}")  # List found XML files
                        xml_files.append(os.path.join(root, file))
                        
    print(f"Debug: Found {len(xml_files)} XML files in search directories.")
    return xml_files


def set_alternating_row_colors(ws, rows):
    """
    Set alternating row colors in the worksheet.

    Args:
        ws (Worksheet): Excel worksheet object.
        rows (list): List of rows to apply alternating colors to.
    """
    medium_light_grey_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=len(rows) + 1), start=2):
        fill_color = medium_light_grey_fill if i % 2 == 0 else white_fill
        for cell in row:
            cell.fill = fill_color
            cell.border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin')) 

def write_to_excel(fields, output_file):
    """
    Write field information to an Excel file.

    Args:
        fields (list): List of dictionaries containing field information.
        output_file (str): Path to the output Excel file.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "All Objects"
        headers = list(COLUMN_HEADINGS.values())
        set_headers_and_widths(ws, headers, fields)
        set_alternating_row_colors(ws, fields)
        write_fields_to_sheet(ws, fields, headers)
        
        # Freeze the first row
        ws.freeze_panes = 'A2'
        
        wb.save(output_file)
        print(f'Excel file generated successfully: {output_file}')
    except Exception as e:
        print(f"Error writing Excel file '{output_file}': {e}")


def set_headers_and_widths(ws, headers, fields):
    """
    Set column headers and widths in the worksheet.

    Args:
        ws (Worksheet): Excel worksheet object.
        headers (list): List of column headers.
        fields (list): List of dictionaries containing field information.
    """
    header_fill = PatternFill(start_color='03584E', end_color='03584E', fill_type='solid')
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True, name='Spartan Light', size=13, color='F6AB93')
        cell.fill = header_fill
        column_values = [str(field.get(header, "")) for field in fields]
        column_width = max(len(header), max(len(value) for value in column_values)) if column_values else len(header)
        ws.column_dimensions[col_letter].width = column_width + 2

def write_fields_to_sheet(ws, fields, headers):
    """
    Write field information to the worksheet.

    Args:
        ws (Worksheet): Excel worksheet object.
        fields (list): List of dictionaries containing field information.
        headers (list): List of column headers.
    """
    body_font = Font(name='Spartan Light', size=12)  # Assuming you want Spartan Light here too
    for row_num, field in enumerate(fields, 2):
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            value = field.get(header, "")
            if header == 'Field Type' and field.get('Formula'):
                value = f"Formula: {value}"
            elif isinstance(value, str) and value.startswith('='):
                value = "'" + value
            cell.value = value
            cell.font = body_font


def main():
    # Setup argument parser
    parser = argparse.ArgumentParser(description='Generate Excel data dictionary from Salesforce field metadata XML files.')
    parser.add_argument('-fields', choices=['all', 'standard', 'custom'], default='all',
                        help='Specify which fields to include: all, standard, or custom.')
    args = parser.parse_args()

    print(f"Fields argument value: {args.fields}")

    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_folder = os.path.join(script_dir, 'DataDictionary')
        shutil.rmtree(output_folder, ignore_errors=True)
        os.makedirs(output_folder, exist_ok=True)
        
        # Find XML files
        xml_files = find_xml_files()
        
        # Parse XML files and filter fields based on the -fields argument
        fields = []
        for xml_file in xml_files:
            field_info = parse_xml(xml_file, args.fields)
            if field_info is not None:  # Only add if field_info is not None
                fields.append(field_info)
        
        # Sort the fields list by Object name and then by API Name within each object
        fields_sorted = sorted(fields, key=lambda x: (x['Object'], x['API Name']))
        
        # Write to Excel
        write_to_excel(fields_sorted, os.path.join('DataDictionary.xlsx'))
        print('Conversion completed successfully.')
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
